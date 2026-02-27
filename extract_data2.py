# -*- coding: utf-8 -*-
"""
2024 하반기 + 본부별 데이터 + 2026-01 데이터 추출
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from collections import defaultdict

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. 2024년 하반기 + 연간 합계
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("=" * 60)
print("2024년 월별 취급고 전체")
print("=" * 60)

path_2024 = r'C:\Users\Administrator\Downloads\나스미디어 2024년 12월 취급고 현황_0116(f) (2).xlsx'
wb24 = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)
ws24 = wb24['2024년 연간 결산']

# 행별로 스캔하여 월별 취급고 행 찾기
all_rows = list(ws24.iter_rows(values_only=True))
wb24.close()

# 월별 데이터 추출
data_2024 = {}
data_2023 = {}
months_2024 = []
months_2023 = []

for i, row in enumerate(all_rows):
    # 행 40: 1. 2024년 월별 취급고
    # 행 41: 구분, 1월, 2월, ..., 상반기 합계
    # 행 42: 2023년, ...
    # 행 43: 2024년, ...
    # 행 46: 구분, 7월, 8월, ..., 하반기 합계
    # 행 47: 2023년, ...
    # 행 48: 2024년, ...
    vals = list(row)
    if vals[1] == '2024년' and vals[2] and isinstance(vals[2], (int, float)):
        # 월별 데이터 행
        print(f"행{i+1} (2024년): {[v for v in vals[:14] if v is not None]}")
    if vals[1] == '2023년' and vals[2] and isinstance(vals[2], (int, float)):
        print(f"행{i+1} (2023년 in 2024file): {[v for v in vals[:14] if v is not None]}")

# 더 상세한 데이터: 전체 비어있지 않은 행 보기
print("\n--- 전체 비어있지 않은 행 (40~60행) ---")
for i, row in enumerate(all_rows[38:65], start=39):
    vals = [v for v in list(row)[:16] if v is not None]
    if vals:
        print(f"  행{i}: {vals[:10]}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. 2024 본부별 월별 데이터 (4-1 피벗 시트)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "=" * 60)
print("2024년 본부별 취급고 (4-1. 피벗 시트)")
print("=" * 60)

wb24b = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)
if '4-1. 피벗' in wb24b.sheetnames:
    ws_pivot = wb24b['4-1. 피벗']
    print(f"크기: {ws_pivot.max_row}행 x {ws_pivot.max_column}열")
    pivot_rows = list(ws_pivot.iter_rows(max_row=40, values_only=True))
    for i, row in enumerate(pivot_rows):
        if any(v for v in row if v is not None):
            vals = [str(v)[:25] if v else '' for v in row[:16]]
            print(f"  행{i+1}: {[v for v in vals if v][:10]}")
wb24b.close()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. 2024 1. 12월 취급고 요약 시트 – 본부별 데이터
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "=" * 60)
print("2024년 12월 취급고 요약 – 본부별 분석")
print("=" * 60)

wb24c = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)
ws_summary = wb24c['1. 12월 취급고 요약']
print(f"크기: {ws_summary.max_row}행 x {ws_summary.max_column}열")
sum_rows = list(ws_summary.iter_rows(values_only=True))
wb24c.close()

print("비어있지 않은 행:")
for i, row in enumerate(sum_rows):
    vals = [v for v in list(row)[:16] if v is not None]
    if vals and len(vals) > 1:
        print(f"  행{i+1}: {[str(v)[:25] for v in vals[:8]]}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. 2026-01 취급고 – XML 직접 파싱 시도
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "=" * 60)
print("2026-01 취급고 – XML 직접 파싱")
print("=" * 60)

import zipfile
xls_2026 = r'C:\Users\Administrator\Downloads\2026-01~2026-01 취급고_20260128_150943 (2).xls'

# 먼저 파일 헤더 확인
with open(xls_2026, 'rb') as f:
    header = f.read(100)
print(f"파일 헤더: {header[:50]}")

# XML 파일이라면 직접 파싱
if b'<?xml' in header:
    print("→ SpreadsheetML XML 형식 감지")
    try:
        import xml.etree.ElementTree as ET
        # 작은 청크만 파싱 (파일이 크므로)
        tree = ET.parse(xls_2026)
        root = tree.getroot()
        ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

        # 워크시트 목록
        sheets = root.findall('.//ss:Worksheet', ns)
        if not sheets:
            # 다른 네임스페이스 시도
            sheets = root.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Worksheet')

        print(f"시트 수: {len(sheets)}")
        for sheet in sheets[:3]:
            name = sheet.get('{urn:schemas-microsoft-com:office:spreadsheet}Name') or sheet.get('Name', '?')
            print(f"  시트: {name}")

            # 첫 10행 데이터
            rows = sheet.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Row')
            for j, row in enumerate(rows[:10]):
                cells = row.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                vals = []
                for cell in cells[:10]:
                    data = cell.find('{urn:schemas-microsoft-com:office:spreadsheet}Data')
                    if data is not None and data.text:
                        vals.append(data.text[:20])
                if vals:
                    print(f"    행{j+1}: {vals}")
    except Exception as e:
        print(f"XML 파싱 오류: {e}")
        import traceback
        traceback.print_exc()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 5. 인원명부 – 팀별 인원 상세
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "=" * 60)
print("인원명부 – 본부/실/팀별 상세 인원")
print("=" * 60)

roster_path = r'C:\Users\Administrator\Downloads\인원명부정보_260131 (2).xlsx'
wb_r = openpyxl.load_workbook(roster_path, data_only=True, read_only=True)
ws_r = wb_r['Sheet']
rows = list(ws_r.iter_rows(min_row=2, values_only=True))
wb_r.close()

# 본부별 재직자 카운트 (직원구분별)
bonbu_map = defaultdict(lambda: defaultdict(int))
for row in rows:
    if row[7] != '재직':
        continue
    bonbu = row[4] or '-'
    bonbu_map[bonbu]['total'] += 1

print("[본부별 인원 (재직자 기준)]")
for bk in sorted(bonbu_map.keys()):
    print(f"  {bk}: {bonbu_map[bk]['total']}명")

# 미디어본부 상세 (현재 8명인데 실제는 52명)
print("\n[미디어본부 상세]")
for row in rows:
    if row[7] == '재직' and row[4] == '미디어본부':
        print(f"  {row[2]} / {row[5] or '-'} / {row[6] or '-'} / {row[9]} / {row[10]}")

print("\n[JavaScript HEADCOUNT 업데이트 코드]")
hc_map = {
    '광고1본부': bonbu_map.get('광고1본부', {}).get('total', 94),
    '광고2본부': bonbu_map.get('광고2본부', {}).get('total', 62),
    '광고3본부': bonbu_map.get('광고3본부', {}).get('total', 48),
    '미디어본부': bonbu_map.get('미디어본부', {}).get('total', 52),
    '플랫폼사업본부': bonbu_map.get('플랫폼사업본부', {}).get('total', 71),
}
total_hc = sum(hc_map.values())
print(f"const HEADCOUNT={{'광고1본부':{hc_map['광고1본부']},'광고2본부':{hc_map['광고2본부']},'광고3본부':{hc_map['광고3본부']},'미디어본부':{hc_map['미디어본부']},'플랫폼사업본부':{hc_map['플랫폼사업본부']}}};")
print(f"// HC_TOTAL: {total_hc}명")
