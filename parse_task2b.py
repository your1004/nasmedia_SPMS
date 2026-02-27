# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

path_2023 = r'C:\Users\Administrator\Downloads\나스미디어 2023년 12월 취급고 현황_240115 (1).xlsx'
wb23 = openpyxl.load_workbook(path_2023, data_only=True, read_only=True)
ws23 = wb23['2. 월별 취급고 누적']
rows23 = list(ws23.iter_rows(values_only=True))

print("--- Rows 4-25 full data (check col structure) ---")
for i in range(3, 22):
    row = rows23[i]
    vals = [str(v)[:30] if v is not None else 'None' for v in row[:16]]
    print(f"  Row{i+1}: {vals}")

print("\n--- 2023 총 취급고 rows (row 6, idx 5) ---")
total_row_23 = rows23[5]  # row 6 = 2023년
print("Row6:", [str(v)[:20] if v is not None else 'None' for v in total_row_23[:16]])

total_row_22 = rows23[4]  # row 5 = 2022년  
print("Row5:", [str(v)[:20] if v is not None else 'None' for v in total_row_22[:16]])

print("\n--- 2023 본부별 rows (rows 15-18, idx 14-17) ---")
for i in [14, 15, 16, 17]:
    row = rows23[i]
    print(f"  Row{i+1}: {[str(v)[:25] if v is not None else '' for v in row[:14]]}")

wb23.close()

# Also check 1. 12월 취급고 요약
print("\n\n=== 2023 12월 취급고 요약 ===")
wb23b = openpyxl.load_workbook(path_2023, data_only=True, read_only=True)
ws23_dec = wb23b['1. 12월 취급고 요약']
dec_rows23 = list(ws23_dec.iter_rows(values_only=True))
wb23b.close()

print(f"Total rows: {len(dec_rows23)}")
for i, row in enumerate(dec_rows23):
    vals = [v for v in row[:12] if v is not None]
    if vals:
        print(f"  Row{i+1}: {[str(v)[:30] for v in vals[:6]]}")
