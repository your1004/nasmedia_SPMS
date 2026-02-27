# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, xml.etree.ElementTree as ET
from collections import defaultdict
import json

path_2024 = r'C:\Users\Administrator\Downloads\나스미디어 2024년 12월 취급고 현황_0116(f) (2).xlsx'
path_2023 = r'C:\Users\Administrator\Downloads\나스미디어 2023년 12월 취급고 현황_240115 (1).xlsx'
xls_path = r'C:\Users\Administrator\Downloads\2026-01~2026-01 취급고_20260128_150943 (2).xls'

def safe_int(v):
    try: return round(float(v))
    except: return 0

# ==== TASK 1: 2024 ====
wb24 = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)
ws_m = wb24['2. 월별 취급고 누적']
rows24 = list(ws_m.iter_rows(values_only=True))
monthly_2024_total = [safe_int(rows24[5][j]) for j in range(3, 15)]
dept_monthly_2024 = {
    k: [safe_int(rows24[14+i][j]) for j in range(3, 15)]
    for i, k in enumerate(['광고1본부','광고2본부','광고3본부','미디어본부','플랫폼사업본부'])
}
wb24.close()

wb24b = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)
ws_dec = wb24b['1. 12월 취급고 요약']
dec_rows = list(ws_dec.iter_rows(values_only=True))
wb24b.close()

dept_dec_2024 = {}
dec_team_2024 = {}
for i in range(26, 67):
    row = dec_rows[i]
    key = str(row[0]) if row[0] else ''
    name = str(row[1]) if row[1] else ''
    val = safe_int(row[2]) if len(row) > 2 else 0
    if not key: continue
    if key in ('Total','* 대행사 월 기준','* 작성 월 기준'): continue
    if '합계' in key:
        dept_dec_2024[key.replace(' 합계','')] = val
    elif val > 0:
        dec_team_2024[key] = {'name': name, 'gross': val}

top20_adv_dec2024 = []
for i in range(71, 92):
    row = dec_rows[i]
    if not row[0] or not row[1]: continue
    adv = str(row[1])
    val = safe_int(row[2]) if len(row) > 2 else 0
    if val > 0 and adv not in ('대행사월 기준','작성 월 기준 상위'):
        top20_adv_dec2024.append({'광고주': adv, 'gross': val})

# ==== TASK 2: 2023 ====
wb23 = openpyxl.load_workbook(path_2023, data_only=True, read_only=True)
ws23 = wb23['2. 월별 취급고 누적']
rows23 = list(ws23.iter_rows(values_only=True))
monthly_2023_total = [safe_int(rows23[5][j]) for j in range(3, 15)]
dept_monthly_2023 = {
    k: [safe_int(rows23[14+i][j]) for j in range(3, 15)]
    for i, k in enumerate(['광고본부','디지털본부','전략사업본부'])
}
wb23.close()

wb23b = openpyxl.load_workbook(path_2023, data_only=True, read_only=True)
ws23_dec = wb23b['1. 12월 취급고 요약']
dec_rows23 = list(ws23_dec.iter_rows(values_only=True))
wb23b.close()

dept_dec_2023 = {}
for i in range(15, 21):
    row = dec_rows23[i]
    if row[0] and str(row[0]) not in ('구분','Total','* 대행사월 기준'):
        dept_dec_2023[str(row[0])] = safe_int(row[1]) if len(row) > 1 else 0

dec_team_2023 = {}
for i in range(23, 64):
    row = dec_rows23[i]
    key = str(row[0]) if row[0] else ''
    name = str(row[1]) if row[1] else ''
    val = safe_int(row[2]) if len(row) > 2 else 0
    if not key or key in ('구분','Total','* 대행사 월 기준'): continue
    if '합계' in key or '기준' in key: continue
    if val > 0:
        dec_team_2023[key] = {'name': name, 'gross': val}

# ==== TASK 3: 2026-01 ====
ns = 'urn:schemas-microsoft-com:office:spreadsheet'
with open(xls_path, 'r', encoding='utf-8') as f:
    content = f.read()
root = ET.fromstring(content)
ws26 = root.findall(f'.//{{{ns}}}Worksheet')[0]
rows26 = ws26.findall(f'.//{{{ns}}}Row')

def get_row_values(row):
    cells = row.findall(f'{{{ns}}}Cell')
    result = []
    prev_idx = 0
    for cell in cells:
        idx_attr = cell.get(f'{{{ns}}}Index')
        if idx_attr:
            idx = int(idx_attr) - 1
            while prev_idx < idx:
                result.append(None)
                prev_idx += 1
        data_el = cell.find(f'{{{ns}}}Data')
        val = data_el.text if data_el is not None else None
        result.append(val)
        prev_idx = len(result)
    return result

def get_dept_2026(t):
    if '광고1본부' in t: return '광고1본부'
    if '광고2본부' in t: return '광고2본부'
    if '광고3본부' in t: return '광고3본부'
    if any(x in t for x in ['미디어본부','OOH매체','MIC','O&P']): return '미디어본부'
    if any(x in t for x in ['플랫폼사업본부','nap ','TV플랫폼']): return '플랫폼사업본부'
    return '기타'

COL_TEAM=4; COL_CATEGORY=6; COL_ADVERTISER=7; COL_MEDIA=11; COL_GROSS=23

total_gross_2026 = 0.0
by_team_2026 = defaultdict(float)
by_media_2026 = defaultdict(float)
by_adv_2026 = defaultdict(float)
by_cat_2026 = defaultdict(float)
by_dept_2026 = defaultdict(float)

for r_idx in range(2, len(rows26)):
    rv = get_row_values(rows26[r_idx])
    gross_raw = rv[COL_GROSS] if len(rv) > COL_GROSS else None
    if gross_raw is None: continue
    try: gross = float(gross_raw)
    except: continue
    if gross == 0: continue
    team = str(rv[COL_TEAM]) if len(rv) > COL_TEAM and rv[COL_TEAM] else 'unknown'
    cat = str(rv[COL_CATEGORY]) if len(rv) > COL_CATEGORY and rv[COL_CATEGORY] else 'unknown'
    adv = str(rv[COL_ADVERTISER]) if len(rv) > COL_ADVERTISER and rv[COL_ADVERTISER] else 'unknown'
    media = str(rv[COL_MEDIA]) if len(rv) > COL_MEDIA and rv[COL_MEDIA] else 'unknown'
    dept = get_dept_2026(team)
    total_gross_2026 += gross
    by_team_2026[team] += gross
    by_media_2026[media] += gross
    by_adv_2026[adv] += gross
    by_cat_2026[cat] += gross
    by_dept_2026[dept] += gross

def to_sorted_dict(d, top=None):
    items = sorted(d.items(), key=lambda x: -x[1])
    if top: items = items[:top]
    return {k: round(v) for k, v in items}

top20_adv_2026 = [
    {'rank': i+1, 'advertiser': k, 'gross': round(v), 'pct': round(v/total_gross_2026*100, 2)}
    for i, (k, v) in enumerate(sorted(by_adv_2026.items(), key=lambda x: -x[1])[:20])
]

output = {
    "task1_2024": {
        "monthly_total": monthly_2024_total,
        "annual_total": sum(monthly_2024_total),
        "dept_monthly": dept_monthly_2024,
        "dept_annual": {k: sum(v) for k, v in dept_monthly_2024.items()},
        "dec_dept_gross": dept_dec_2024,
        "dec_team_gross": dec_team_2024,
        "dec_top20_advertiser": top20_adv_dec2024,
    },
    "task2_2023": {
        "monthly_total": monthly_2023_total,
        "annual_total": sum(monthly_2023_total),
        "dept_monthly": dept_monthly_2023,
        "dept_annual": {k: sum(v) for k, v in dept_monthly_2023.items()},
        "dec_dept_gross": dept_dec_2023,
        "dec_team_gross": dec_team_2023,
    },
    "task3_2026_jan": {
        "total_gross": round(total_gross_2026),
        "by_dept": to_sorted_dict(by_dept_2026),
        "by_team": to_sorted_dict(by_team_2026),
        "by_media_top30": to_sorted_dict(by_media_2026, 30),
        "by_category": to_sorted_dict(by_cat_2026),
        "top20_advertiser": top20_adv_2026,
    }
}

with open(r'C:\D\Claude Code Project\nasmedia_SPMS\extracted_data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print("JSON written successfully")
print(f"2024 annual total: {sum(monthly_2024_total):,.0f}")
print(f"2023 annual total: {sum(monthly_2023_total):,.0f}")
print(f"2026 Jan gross:    {round(total_gross_2026):,.0f}")
