# -*- coding: utf-8 -*-
"""
실제 취급고/인원명부 데이터 추출 스크립트
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import os

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. 인원명부 – 본부별 인원 수 추출
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("=" * 60)
print("1. 인원명부 분석")
print("=" * 60)

roster_path = r'C:\Users\Administrator\Downloads\인원명부정보_260131 (2).xlsx'
wb_r = openpyxl.load_workbook(roster_path, data_only=True, read_only=True)
ws_r = wb_r['Sheet']

rows = list(ws_r.iter_rows(min_row=2, values_only=True))
wb_r.close()

# 컬럼: No, 사번, 성명, 총괄, 본부, 실, 팀, 재직상태, 성별, 직위, 직책, 직원구분, 메일주소
# 인덱스:  0    1     2     3     4    5   6   7        8     9    10   11      12

print(f"총 인원: {len(rows)}명")
from collections import defaultdict
bonbu_count = defaultdict(int)
bonbu_teams = defaultdict(set)
bonbu_detail = defaultdict(list)
직책_count = defaultdict(int)

for row in rows:
    if not row[7] or row[7] != '재직':  # 재직자만
        continue
    bonbu = row[4] or '-'
    team = row[6] or '-'
    jikchek = row[10] or ''
    bonbu_count[bonbu] += 1
    bonbu_teams[bonbu].add(team)
    직책_count[jikchek] += 1
    bonbu_detail[bonbu].append({
        '성명': row[2], '실': row[5], '팀': team,
        '직위': row[9], '직책': jikchek
    })

print("\n[본부별 재직 인원]")
total_active = 0
for bk in sorted(bonbu_count.keys()):
    cnt = bonbu_count[bk]
    total_active += cnt
    teams = bonbu_teams[bk]
    print(f"  {bk}: {cnt}명 ({len(teams)}팀)")
print(f"  합계: {total_active}명")

print("\n[직책별 인원]")
for k, v in sorted(직책_count.items(), key=lambda x: -x[1])[:10]:
    print(f"  {k}: {v}명")

# 주요 본부 매핑
print("\n[HEADCOUNT 매핑 (주요 5개 본부)]")
# 우리 HTML에서 사용하는 본부명: 광고1본부, 광고2본부, 광고3본부, 미디어본부, 플랫폼사업본부
key_bonbus = {
    '광고1본부': ['광고1본부'],
    '광고2본부': ['광고2본부'],
    '광고3본부': ['광고3본부'],
    '미디어본부': ['미디어본부'],
    '플랫폼사업본부': ['플랫폼사업본부'],
}
headcount = {}
for hk, search_keys in key_bonbus.items():
    cnt = 0
    for bk in bonbu_count.keys():
        for sk in search_keys:
            if sk in str(bk):
                cnt += bonbu_count[bk]
                break
    headcount[hk] = cnt
    print(f"  '{hk}': {cnt}")
print(f"  HC_TOTAL: {sum(headcount.values())}")

# 상세 팀 목록 출력 (광고1본부)
print("\n[광고1본부 팀 상세]")
for p in bonbu_detail.get('광고1본부', [])[:5]:
    print(f"  {p}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. 2024년 취급고 월별 데이터
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "=" * 60)
print("2. 2024년 12월 취급고 현황 – 월별 데이터")
print("=" * 60)

path_2024 = r'C:\Users\Administrator\Downloads\나스미디어 2024년 12월 취급고 현황_0116(f) (2).xlsx'
wb24 = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)

# 시트 목록 확인
print("시트 목록:", wb24.sheetnames)

# '2. 월별 취급고 누적' 시트 탐색
target_sheet = None
for sh in wb24.sheetnames:
    if '월별' in sh or '누적' in sh or '연간' in sh:
        target_sheet = sh
        print(f"\n→ 탐색 대상 시트: '{sh}'")
        break

if target_sheet:
    ws24 = wb24[target_sheet]
    print(f"  크기: {ws24.max_row}행 x {ws24.max_column}열")
    # 처음 20행 출력
    for i, row in enumerate(ws24.iter_rows(max_row=25, values_only=True)):
        if any(v for v in row if v is not None):
            vals = [str(v)[:20] if v else '' for v in row[:15]]
            print(f"  행{i+1}: {vals}")

# '2024년 연간 결산' 시트 탐색
print("\n→ '2024년 연간 결산' 시트")
ws_annual = wb24['2024년 연간 결산']
print(f"  크기: {ws_annual.max_row}행 x {ws_annual.max_column}열")
non_empty_rows = []
for i, row in enumerate(ws_annual.iter_rows(values_only=True)):
    if any(v for v in row if v is not None):
        non_empty_rows.append((i+1, row[:20]))

print(f"  비어있지 않은 행 수: {len(non_empty_rows)}")
for rn, row in non_empty_rows[:30]:
    vals = [str(v)[:20] if v else '' for v in row]
    if any(v for v in vals):
        print(f"  행{rn}: {[v for v in vals if v][:10]}")

wb24.close()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. 2026-01 취급고 (xls -> xml 형식)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "=" * 60)
print("3. 2026-01 취급고 파일")
print("=" * 60)

xls_2026 = r'C:\Users\Administrator\Downloads\2026-01~2026-01 취급고_20260128_150943 (2).xls'
import shutil

# xml 형식의 xls이므로 xlsx로 복사해서 읽기
tmp_xlsx = r'C:\Users\Administrator\Downloads\tmp_2026_01.xlsx'
try:
    shutil.copy2(xls_2026, tmp_xlsx)
    wb26 = openpyxl.load_workbook(tmp_xlsx, data_only=True, read_only=True)
    print(f"시트 목록: {wb26.sheetnames}")
    for sh_name in wb26.sheetnames[:3]:
        ws26 = wb26[sh_name]
        print(f"\n시트: {sh_name} ({ws26.max_row}행 x {ws26.max_column}열)")
        cnt = 0
        for row in ws26.iter_rows(max_row=20, values_only=True):
            if any(v for v in row if v is not None):
                print(f"  {[str(v)[:20] if v else '' for v in row[:12]]}")
                cnt += 1
    wb26.close()
except Exception as e:
    print(f"오류: {e}")
finally:
    if os.path.exists(tmp_xlsx):
        os.remove(tmp_xlsx)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. 2023년 취급고 – 연간 합계 확인
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
print("\n" + "=" * 60)
print("4. 2023년 취급고 연간 결산")
print("=" * 60)

path_2023 = r'C:\Users\Administrator\Downloads\나스미디어 2023년 12월 취급고 현황_240115 (1).xlsx'
wb23 = openpyxl.load_workbook(path_2023, data_only=True, read_only=True)
print("시트 목록:", wb23.sheetnames)

ws23_annual = wb23['2023년 연간 결산']
non_empty = []
for i, row in enumerate(ws23_annual.iter_rows(values_only=True)):
    if any(v for v in row if v is not None):
        non_empty.append((i+1, row[:20]))

for rn, row in non_empty[:30]:
    vals = [str(v)[:25] if v else '' for v in row]
    if any(v for v in vals):
        print(f"  행{rn}: {[v for v in vals if v][:8]}")

wb23.close()
