# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict

# ============================================================
# TASK 2: 2023 월별 취급고
# ============================================================
print("="*60)
print("TASK 2: 2023 Monthly Revenue")
print("="*60)

path_2023 = r'C:\Users\Administrator\Downloads\나스미디어 2023년 12월 취급고 현황_240115 (1).xlsx'
wb23 = openpyxl.load_workbook(path_2023, data_only=True, read_only=True)
print("2023 sheets:", wb23.sheetnames)

# --- Sheet: 2. 월별 취급고 누적 ---
print("\n--- Sheet: '2. 월별 취급고 누적' ---")
ws23 = wb23['2. 월별 취급고 누적']
rows23 = list(ws23.iter_rows(values_only=True))
print(f"Size: {len(rows23)} rows")
for i, row in enumerate(rows23[:35]):
    if any(v for v in row if v is not None):
        vals = [str(v)[:28] if v is not None else '' for v in row[:16]]
        nv = [v for v in vals if v]
        if nv:
            print(f"  Row{i+1}: {nv[:8]}")

wb23.close()
