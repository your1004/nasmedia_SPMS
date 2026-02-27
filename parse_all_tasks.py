# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, json, re, xml.etree.ElementTree as ET
from collections import defaultdict

# ============================================================
# TASK 1: 2024 월별 취급고 - 전체 팀 데이터 + 12월 요약
# ============================================================
print("="*60)
print("TASK 1a: 2024 월별 취급고 누적 - 모든 팀")
print("="*60)

path_2024 = r'C:\Users\Administrator\Downloads\나스미디어 2024년 12월 취급고 현황_0116(f) (2).xlsx'
wb24 = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)
ws_monthly = wb24['2. 월별 취급고 누적']
monthly_rows = list(ws_monthly.iter_rows(values_only=True))

# Row 6 = 2024 monthly totals (row index 5)
# Row 15-21 = department monthly data (rows index 14-20)
# Row 25+ = team monthly data

months = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']

# Total 2024 monthly (row 6, cols 3-14)
total_row = monthly_rows[5]  # row index 5 = row 6
monthly_2024_total = []
for j in range(3, 15):
    v = total_row[j]
    monthly_2024_total.append(round(float(v)) if v is not None else 0)
print("2024 Total monthly:", monthly_2024_total)

# Department monthly (rows 15-20)
dept_rows = {
    '광고1본부': monthly_rows[14],
    '광고2본부': monthly_rows[15],
    '광고3본부': monthly_rows[16],
    '미디어본부': monthly_rows[17],
    '플랫폼사업본부': monthly_rows[18],
}
dept_monthly_2024 = {}
for dept, row in dept_rows.items():
    vals = []
    for j in range(3, 15):
        v = row[j]
        vals.append(round(float(v)) if v is not None else 0)
    dept_monthly_2024[dept] = vals
    print(f"{dept}: {vals}")

# Team monthly data (rows 26+, until we hit a non-team row)
print("\n--- Team Monthly 2024 ---")
team_monthly_2024 = {}
for i in range(25, len(monthly_rows)):
    row = monthly_rows[i]
    if row[1] is None or not str(row[1]).strip():
        break
    team_key = str(row[1]) if row[1] else ''
    team_name = str(row[2]) if row[2] else ''
    if not team_key or '본부' not in team_key:
        continue
    vals = []
    for j in range(3, 15):
        v = row[j]
        vals.append(round(float(v)) if v is not None else 0)
    team_monthly_2024[f"{team_key}({team_name})"] = vals

print(f"Team count: {len(team_monthly_2024)}")
for k, v in list(team_monthly_2024.items())[:5]:
    print(f"  {k}: total={sum(v):,.0f}")

# Find all teams in the sheet
print("\n--- All teams detected ---")
for i in range(25, 100):
    if i >= len(monthly_rows):
        break
    row = monthly_rows[i]
    if row[1] is not None:
        print(f"  Row{i+1}: [{row[1]}] [{row[2]}] col3={row[3]}")

wb24.close()

# ============================================================
# TASK 1b: 2024 - Sheet "1. 12월 취급고 요약"
# ============================================================
print("\n" + "="*60)
print("TASK 1b: 2024년 12월 취급고 요약")
print("="*60)

wb24b = openpyxl.load_workbook(path_2024, data_only=True, read_only=True)
ws_dec = wb24b['1. 12월 취급고 요약']
dec_rows = list(ws_dec.iter_rows(values_only=True))
wb24b.close()

print(f"Total rows: {len(dec_rows)}")
for i, row in enumerate(dec_rows):
    vals = [v for v in row[:16] if v is not None]
    if vals:
        str_vals = [str(v)[:30] for v in vals[:8]]
        print(f"  Row{i+1}: {str_vals}")
