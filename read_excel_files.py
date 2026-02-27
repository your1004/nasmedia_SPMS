# -*- coding: utf-8 -*-
"""
암호해제된 취급고/인원명부 파일 읽기 및 데이터 추출
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os

files = [
    r'C:\Users\Administrator\Downloads\나스미디어 2024년 12월 취급고 현황_0116(f) (2).xlsx',
    r'C:\Users\Administrator\Downloads\인원명부정보_260131 (2).xlsx',
    r'C:\Users\Administrator\Downloads\나스미디어 2023년 12월 취급고 현황_240115 (1).xlsx',
    r'C:\Users\Administrator\Downloads\2026-01~2026-01 취급고_20260128_150943 (2).xls',
]

for f in files:
    exists = os.path.exists(f)
    size = os.path.getsize(f) if exists else 0
    print(f"{'존재' if exists else '없음'}: {os.path.basename(f)} ({size:,} bytes)")

print("\n=== openpyxl로 읽기 시도 ===")
try:
    import openpyxl
    print("openpyxl 버전:", openpyxl.__version__)
except ImportError:
    print("openpyxl 없음 - pip install openpyxl")

# 인원명부 파일 읽기
print("\n--- 인원명부정보_260131 (2).xlsx ---")
roster_path = r'C:\Users\Administrator\Downloads\인원명부정보_260131 (2).xlsx'
if os.path.exists(roster_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(roster_path, data_only=True, read_only=True)
        print(f"시트 목록: {wb.sheetnames}")
        for sh_name in wb.sheetnames[:3]:
            ws = wb[sh_name]
            print(f"\n시트: {sh_name} ({ws.max_row}행 x {ws.max_column}열)")
            rows = list(ws.iter_rows(max_row=5, values_only=True))
            for r in rows:
                print(r)
        wb.close()
    except Exception as e:
        print(f"openpyxl 오류: {e}")
        # xlrd 시도
        try:
            import xlrd
            wb2 = xlrd.open_workbook(roster_path)
            print(f"xlrd 시트: {wb2.sheet_names()}")
        except Exception as e2:
            print(f"xlrd 오류: {e2}")

# 2026-01 취급고 파일 읽기
print("\n--- 2026-01~2026-01 취급고 파일 ---")
xls_path = r'C:\Users\Administrator\Downloads\2026-01~2026-01 취급고_20260128_150943 (2).xls'
if os.path.exists(xls_path):
    try:
        import xlrd
        wb = xlrd.open_workbook(xls_path)
        print(f"시트 목록: {wb.sheet_names()}")
        for sh_name in wb.sheet_names()[:3]:
            ws = wb.sheet_by_name(sh_name)
            print(f"\n시트: {sh_name} ({ws.nrows}행 x {ws.ncols}열)")
            for r in range(min(5, ws.nrows)):
                print(list(ws.row_values(r))[:10])
    except Exception as e:
        print(f"xlrd 오류: {e}")
        # openpyxl 시도
        try:
            import openpyxl
            wb2 = openpyxl.load_workbook(xls_path, data_only=True, read_only=True)
            print(f"openpyxl 시트: {wb2.sheetnames}")
        except Exception as e2:
            print(f"openpyxl 오류: {e2}")

# 2024년 12월 취급고 파일
print("\n--- 나스미디어 2024년 12월 취급고 현황 ---")
xlsx_2024 = r'C:\Users\Administrator\Downloads\나스미디어 2024년 12월 취급고 현황_0116(f) (2).xlsx'
if os.path.exists(xlsx_2024):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_2024, data_only=True, read_only=True)
        print(f"시트 목록: {wb.sheetnames}")
        for sh_name in wb.sheetnames[:2]:
            ws = wb[sh_name]
            print(f"\n시트: {sh_name} ({ws.max_row}행 x {ws.max_column}열)")
            rows = list(ws.iter_rows(max_row=8, values_only=True))
            for r in rows:
                print(r[:10])
        wb.close()
    except Exception as e:
        print(f"openpyxl 오류: {e}")

# 2023년 12월 취급고 파일
print("\n--- 나스미디어 2023년 12월 취급고 현황 ---")
xlsx_2023 = r'C:\Users\Administrator\Downloads\나스미디어 2023년 12월 취급고 현황_240115 (1).xlsx'
if os.path.exists(xlsx_2023):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_2023, data_only=True, read_only=True)
        print(f"시트 목록: {wb.sheetnames}")
        for sh_name in wb.sheetnames[:2]:
            ws = wb[sh_name]
            print(f"\n시트: {sh_name} ({ws.max_row}행 x {ws.max_column}열)")
            rows = list(ws.iter_rows(max_row=8, values_only=True))
            for r in rows:
                print(r[:10])
        wb.close()
    except Exception as e:
        print(f"openpyxl 오류: {e}")
